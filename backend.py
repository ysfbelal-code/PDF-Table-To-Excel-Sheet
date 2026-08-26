import os
import re
import time
from pathlib import Path
import pandas as pd

from img2table.document.pdf import PDF

# ==========================================
# 0. PATH & COMPAT FIXES (supports PyInstaller frozen exe)
# ==========================================
import sys
if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
    _BASE = Path(sys._MEIPASS)
    _SCRIPT_DIR = Path(__file__).parent.resolve() if "__file__" in globals() else _BASE
    # In frozen exe, Tesseract/Poppler are bundled as datas in _MEIPASS
    _TESS_PATH = _BASE / "Tesseract-OCR"
    _POPPLER_PATH = _BASE / "poppler-26.02.0" / "Library" / "bin"
else:
    _BASE = Path(__file__).parent.resolve()
    _SCRIPT_DIR = _BASE
    _TESS_PATH = _SCRIPT_DIR / "Tesseract-OCR"
    _POPPLER_PATH = _SCRIPT_DIR / "poppler-26.02.0" / "Library" / "bin"

for _p in [_TESS_PATH, _POPPLER_PATH]:
    if _p.exists() and str(_p) not in os.environ["PATH"]:
        os.environ["PATH"] += os.pathsep + str(_p)
# Also ensure TESSDATA_PREFIX points to bundled tessdata when frozen
if (_TESS_PATH / "tessdata").exists():
    os.environ["TESSDATA_PREFIX"] = str(_TESS_PATH / "tessdata")
elif (_TESS_PATH).exists():
    # Some builds keep eng.traineddata directly in Tesseract-OCR
    if any(_TESS_PATH.glob("*.traineddata")):
        os.environ["TESSDATA_PREFIX"] = str(_TESS_PATH)

# img2table 2.0 bug: OCRData._group_words_by_parent crashes on None values (Tesseract only)
try:
    from collections import defaultdict as _dd
    from img2table.ocr._types import OCRData as _OCRData
    def _patched_group(words):
        parent_words = _dd(list)
        for w in words:
            if w.get("value") is None:
                continue
            parent_words[w.get("parent")].append(w)
        if not parent_words:
            return []
        lines = [{"x1": min(x["x1"] for x in words_line), "y1": min(x["y1"] for x in words_line),
                  "value": " ".join([str(x["value"]) for x in sorted(words_line, key=lambda wd: wd["x1"])]).strip()}
                 for words_line in parent_words.values()]
        return [line["value"] for line in sorted(lines, key=lambda line: (line["y1"], line["x1"]))]
    _OCRData._group_words_by_parent = staticmethod(_patched_group)
except Exception:
    pass

INPUT_UNKNOWN_PDF = "file.pdf"
FINAL_EXCEL_OUTPUT = "fully_dynamic_extraction.xlsx"


def _sanitize_for_excel(val):
    """Remove invalid XML characters that break Excel files.
    Valid XML 1.0 char ranges: #x9 | #xA | #xD | [#x20-#xD7FF] | [#xE000-#xFFFD] | [#x10000-#x10FFFF]
    """
    if val is None:
        return None
    s = str(val)
    # Remove control characters except tab (0x09), newline (0x0A), carriage return (0x0D)
    # Also remove null bytes and other problematic characters
    cleaned = "".join(
        ch for ch in s
        if ord(ch) == 0x09 or ord(ch) == 0x0A or ord(ch) == 0x0D
        or (0x20 <= ord(ch) <= 0xD7FF)
        or (0xE000 <= ord(ch) <= 0xFFFD)
        or (ord(ch) > 0xFFFF)  # Allow supplementary planes
    )
    return cleaned.strip() if cleaned.strip() else None


def _get_ocr(multilingual=True):
    """Try DocTR (best for this scan: 95 words avg 81 vs Tesseract 69 avg 35), fallback to EasyOCR, then Tesseract.
    multilingual=True enables English+Arabic in same cell (user requirement)."""
    # 1) DocTR — verified 13x5 perfect table on file.pdf (474x333 JPEG, blue grid)
    # detect_language=True enables Arabic+English merged cells (كشف حساب / Statement)
    try:
        from img2table.ocr import DocTR
        # detect_language True costs ~10% speed but required for bilingual
        ocr = DocTR(detect_language=multilingual)
        print(f"[*] OCR engine: DocTR (doctr, detect_language={multilingual})")
        return ocr
    except Exception as e:
        print(f"[!] DocTR unavailable: {e}")

    # 2) EasyOCR — 40 words, better than Tesseract but truncated bbox
    # For Arabic+English merged, need lang=["en","ar"] (user: English and Arabic have to be in same cell)
    try:
        from img2table.ocr import EasyOCR
        langs = ["en", "ar"] if multilingual else ["en"]
        ocr = EasyOCR(lang=langs, kw={"gpu": False, "verbose": False})
        print(f"[*] OCR engine: EasyOCR (fallback, langs={langs})")
        return ocr
    except Exception as e:
        print(f"[!] EasyOCR unavailable: {e}")

    # 3) Tesseract — last resort, needs PATH fix and psm=6
    # Note: Tesseract-OCR/tessdata only has eng.traineddata; ara would need download, so bilingual will be English-only
    try:
        from img2table.ocr import TesseractOCR
        ocr = TesseractOCR(n_threads=1, lang="eng", psm=6)
        print("[*] OCR engine: Tesseract (fallback, psm=6, English-only)")
        if multilingual:
            print("[!] WARNING: Tesseract fallback cannot handle Arabic (missing ara.traineddata); results will be English-only")
        return ocr
    except OSError as e:
        raise OSError(f"No OCR engine available. Tesseract check: {_TESS_PATH} exists={_TESS_PATH.exists()}. {e}") from e


def _clean_cell(val):
    """Post-OCR cleanup for GENERAL JOURNAL quirks (DocTR outputs S for $, CI for C1, etc.)."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() == "none":
        return None
    s = re.sub(r"\s*\n\s*", " ", s)
    s = s.strip()
    # Apply Excel sanitization
    return _sanitize_for_excel(s)


def _looks_like_statement(df):
    """Detect statement-like tables: sparse right-side numeric column + dense text column.
    Returns True only if structure matches (e.g., bank statement with Balance column).
    Curriculum/syllabus tables do NOT match this pattern.
    """
    if df.empty or df.shape[0] < 3 or df.shape[1] < 3:
        return False
    counts = {c: df[c].notna().sum() for c in df.columns}
    total = len(df)
    # Need at least one sparse column (5-50% filled)
    sparse_cols = [c for c, cnt in counts.items() if 1 < cnt < total * 0.5]
    if not sparse_cols:
        return False
    # Need at least one dense column (>= 80% filled)
    dense_cols = [c for c, cnt in counts.items() if cnt >= total * 0.8]
    if not dense_cols:
        return False
    # Sparse column must be numeric (date/amount/ID pattern)
    amount_re = re.compile(r"\d+[\.,]\d{2}|\d{4,}|^\d+[-/]\d+[-/]\d+")
    for c in sparse_cols:
        col_vals = [str(v) for v in df[c] if v is not None and str(v).strip()]
        if col_vals and sum(1 for v in col_vals if amount_re.search(v)) / len(col_vals) > 0.5:
            return True
    return False


def _merge_multiline_generic(df):
    """Merge vertically split multi-line values into first cell (user request).
    Only applies to statement-like tables (sparse right-side numeric column + dense text column).
    For non-statement tables (curriculum, syllabus, schedules), returns df unchanged.
    """
    if df.empty or df.shape[0] < 3:
        return df
    # NEW: structural guard — only merge if table looks like a statement
    if not _looks_like_statement(df):
        return df
    try:
        # Count non-None per column
        counts = {c: df[c].notna().sum() for c in df.columns}
        total = len(df)
        anchor_col = None
        min_count = total
        for c, cnt in counts.items():
            if 1 < cnt < total * 0.5 and cnt < min_count:
                min_count = cnt
                anchor_col = c
        if anchor_col is None:
            return df
        desc_col = max(counts, key=lambda k: counts[k])
        if counts[desc_col] < total * 0.8:
            return df
        if anchor_col == desc_col:
            return df
        anchor_indices = [i for i, r in df.iterrows() if pd.notna(r[anchor_col]) and str(r[anchor_col]).strip() != ""]
        if len(anchor_indices) < 2:
            return df
        date_re = re.compile(r"^\d{2}[-/]\d{2}[-/]\d{2,4}$")
        for i, row in df.iterrows():
            if i in anchor_indices:
                continue
            d = row[desc_col]
            if pd.isna(d) or not str(d).strip():
                continue
            s = str(d).strip()
            if date_re.match(s):
                continue
        merged_rows = []
        for a in sorted(anchor_indices):
            row = df.iloc[a].copy()
            assigned = [a] + [i for i in range(len(df)) if i not in anchor_indices and min(anchor_indices, key=lambda x: abs(i - x)) == a and not date_re.match(str(df.iloc[i][desc_col]).strip())]
            all_descs = []
            for idx in sorted(set(assigned)):
                d = df.iloc[idx][desc_col]
                if pd.notna(d) and str(d).strip() and not date_re.match(str(d).strip()):
                    all_descs.append(str(d).strip())
            uniq = []
            seen = set()
            for d in all_descs:
                if d not in seen:
                    uniq.append(d)
                    seen.add(d)
            if uniq:
                row[desc_col] = "\n".join(uniq)
            merged_rows.append(row)
        if len(merged_rows) >= 2 and len(merged_rows) < len(df) * 0.8:
            return pd.DataFrame(merged_rows, columns=df.columns)
    except Exception:
        pass
    return df


def _combine_tables(tables_dict):
    """Combine tables with same headers (user request).
    - Groups by column header tuple
    - For near-duplicate tables (only one cell changed), keep only first
    - Concatenates rows for same-header groups into one DataFrame per group
    Returns new dict with combined sheets.
    Dynamic, no hardcoded headers.
    """
    from collections import defaultdict
    # Group by header signature
    groups = defaultdict(list)
    for name, df in tables_dict.items():
        # Use column tuple as key (normalized: strip, lower for comparison? keep original for output)
        key = tuple(str(c).strip() for c in df.columns)
        groups[key].append((name, df))

    combined = {}
    for key, items in groups.items():
        # Deduplicate near-duplicate tables: only one cell difference
        unique_dfs = []
        for name, df in items:
            is_dup = False
            for _, existing in unique_dfs:
                if df.shape != existing.shape:
                    continue
                if list(df.columns) != list(existing.columns):
                    continue
                # Count differing cells (ignoring NaN equality)
                diff = 0
                for r in range(len(df)):
                    for c in range(len(df.columns)):
                        v1 = df.iloc[r, c]
                        v2 = existing.iloc[r, c]
                        # Normalize None/NaN
                        if pd.isna(v1) and pd.isna(v2):
                            continue
                        if str(v1) != str(v2):
                            diff += 1
                            if diff > 1:
                                break
                    if diff > 1:
                        break
                if diff <= 1:
                    # Consider duplicate (only one cell changed) -> keep first, skip this
                    print(f"    -> Deduped near-duplicate table {name} (only {diff} cell(s) differ from kept table)")
                    is_dup = True
                    break
            if not is_dup:
                unique_dfs.append((name, df))

        if len(unique_dfs) == 1:
            # Only one unique table for this header, keep as is but rename to Combined_*
            name, df = unique_dfs[0]
            # Keep original name if only one group total, else use combined name
            if len(groups) == 1:
                combined["Combined"] = df
            else:
                combined[name] = df
        else:
            # Multiple tables with same header -> concatenate vertically
            dfs_to_concat = [df for _, df in unique_dfs]
            # Ensure columns align
            combined_df = pd.concat(dfs_to_concat, ignore_index=True)
            # Create sheet name based on header hash or first table name
            base_name = unique_dfs[0][0].split("_Table_")[0]  # e.g., Page_1
            # Use generic combined name
            sheet_name = f"Combined_{len(combined)+1}"[:31]
            # If all tables share same header and it's the only group, use "Combined"
            if len(groups) == 1:
                sheet_name = "Combined"
            combined[sheet_name] = combined_df
            print(f"    -> Combined {len(unique_dfs)} tables into {sheet_name} ({len(combined_df)} rows, {len(key)} cols)")

    # If only one combined sheet and original had multiple groups, keep them, else if combined produced single sheet, return it
    if len(combined) == 1 and len(tables_dict) > 1:
        # Already combined into one, return
        return combined
    # If combined has same number as original groups, return combined (deduplicated)
    return combined if combined else tables_dict


def enterprise_grade_table_miner(input_path, output_excel, combine_mode="separate"):
    """Extract tables from PDF (scanned or digital) to Excel using DocTR stack.
    combine_mode: 'separate' = each table as separate sheet (default), 'combined' = merge same-header tables and dedupe near-duplicates
    """
    print(f"[*] Analyzing document structural matrices on: {input_path}")
    _input = Path(input_path)
    if not _input.is_absolute() and not _input.exists():
        _alt = _SCRIPT_DIR / _input
        if _alt.exists():
            input_path = str(_alt)
    _output = Path(output_excel)
    if not _output.is_absolute() and Path.cwd().resolve() != _SCRIPT_DIR.resolve():
        output_excel = str(_SCRIPT_DIR / _output)

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Missing target file: {input_path} (resolved against {_SCRIPT_DIR})")

    start_time = time.time()

    ocr = _get_ocr(multilingual=True)

    # Per-page text layer detection (accuracy over speed: handle mixed scanned+digital PDFs)
    # Global sum fails for mixed docs (1 scanned 0 chars + 5 digital 12k chars -> global True but scanned page needs OCR)
    import pypdfium2
    try:
        _doc = pypdfium2.PdfDocument(input_path)
        num_pages = len(_doc)
        page_has_text = {}
        for pi in range(num_pages):
            try:
                cnt = _doc[pi].get_textpage().count_chars()
                # Also check for Arabic/English mix: if count > 50 and not just whitespace, consider text
                page_has_text[pi] = cnt > 50
            except Exception:
                page_has_text[pi] = False
        _doc.close()
        # Log per-page
        for pi, has in page_has_text.items():
            print(f"[*] Page {pi+1}: {'digital (vector text)' if has else 'scanned (image)'} -> {'PdfOCR' if has else 'DocTR/EasyOCR'}")
    except Exception as e:
        print(f"[!] PDF text detection failed: {e}, defaulting to OCR for all pages")
        import pypdfium2 as _p2
        _d = _p2.PdfDocument(input_path)
        num_pages = len(_d)
        _d.close()
        page_has_text = {i: False for i in range(num_pages)}

    # Accuracy-first cascade: try multiple table params and pick best by non-None cells (no speed compromise)
    # For scanned pages, borderless_tables=True mandatory (False->0 tables), but try both for digital
    def _extract_for_page(page_idx, has_text_flag):
        # Returns dict {page_idx: [tables]} with best score
        use_ocr = None if has_text_flag else ocr
        # pdf_text_extraction True for digital (uses PdfOCR, no OCR cost), False for scanned
        doc_page = PDF(src=input_path, pages=[page_idx], pdf_text_extraction=has_text_flag)
        best = None
        best_score = -1
        best_params = None
        # Cascade params: accuracy over speed, try all combos for scanned, limited for digital
        borderless_opts = [True, False] if has_text_flag else [True]  # scanned: only True matters per test_grid_search
        implicit_opts = [True, False]
        mc_opts = [30] if has_text_flag else [50, 30]
        for bl in borderless_opts:
            for ir in implicit_opts:
                for mc in mc_opts:
                    try:
                        res = doc_page.extract_tables(ocr=use_ocr, implicit_rows=ir, borderless_tables=bl, min_confidence=mc)
                        # res is {page_idx: [ExtractedTable]}
                        tables_list = res.get(page_idx, [])
                        if not tables_list:
                            continue
                        # Score by total non-None cells and avg confidence proxy (cell count)
                        score = 0
                        for t in tables_list:
                            df_tmp = t.df
                            # Count non-None after stripping
                            score += df_tmp.notna().sum().sum()
                            # Bonus for reasonable shape (not 1xN)
                            if t.df.shape[0] >= 2 and t.df.shape[1] >= 2:
                                score += 10
                        if score > best_score:
                            best_score = score
                            best = res
                            best_params = (bl, ir, mc)
                    except Exception as ce:
                        continue
        if best is None:
            # Fallback single try
            doc_fallback = PDF(src=input_path, pages=[page_idx], pdf_text_extraction=has_text_flag)
            best = doc_fallback.extract_tables(ocr=use_ocr, implicit_rows=True, borderless_tables=True, min_confidence=50 if not has_text_flag else 30)
            best_params = (True, True, 50 if not has_text_flag else 30)
        if best_params:
            print(f"    -> best params borderless={best_params[0]} implicit_rows={best_params[1]} min_conf={best_params[2]} score={best_score}")
        return best

    print("[*] Computing visual column pathways and spatial grids (per-page cascade)...")
    extracted_tables = {}
    for pi in range(num_pages):
        has = page_has_text.get(pi, False)
        res = _extract_for_page(pi, has)
        if res:
            extracted_tables.update(res)

    tables_to_write = {}
    sheet_counter = 0

    def _is_journal(table, df_raw):
        # Dynamic detection: journal has 5 columns, 10+ rows, many rows with amount in last two cols
        # No hardcoded business names - use structural patterns only
        if df_raw.empty or df_raw.shape[1] != 5 or df_raw.shape[0] < 10:
            return False
        try:
            amount_re = re.compile(r"\d+[\.,]\d{2}")
            amount_rows = 0
            for _, row in df_raw.iterrows():
                last_two = [str(v) for v in row.tolist()[-2:] if v]
                if any(amount_re.search(str(v)) for v in last_two):
                    amount_rows += 1
            if amount_rows / len(df_raw) > 0.3:
                doc_re = re.compile(r"^[A-Za-z]{1,2}\d{1,2}$")
                doc_hits = 0
                for _, row in df_raw.iterrows():
                    mid = str(row.iloc[2]) if len(row) > 2 else ""
                    if doc_re.match(mid.strip()):
                        doc_hits += 1
                if doc_hits >= 2:
                    return True
        except Exception:
            pass
        return False

    for page_num, tables in extracted_tables.items():
        for t_idx, table in enumerate(tables):
            sheet_counter += 1
            raw_matrix = table.content
            cleaned_rows = []
            for row_idx in sorted(raw_matrix.keys()):
                row_cells = [_clean_cell(c.value) for c in raw_matrix[row_idx]]
                cleaned_rows.append(row_cells)

            df = pd.DataFrame(cleaned_rows)

            # Generic cleanup - preserves Arabic+English merged cells (no splitting)
            df = df.map(lambda x: re.sub(r"\s+", " ", str(x)).strip() if x is not None and str(x).lower() != "none" else None)
            df.replace(["", "None", "none"], None, inplace=True)
            df.dropna(how="all", axis=0, inplace=True)
            df.dropna(how="all", axis=1, inplace=True)

            if df.empty:
                continue

            # Branch: Journal-specific vs Generic (bilingual statement)
            if _is_journal(table, df):
                # === JOURNAL PATH (file.pdf) ===
                if df.shape[1] < 5:
                    continue
                # Keep first 5 cols, huge ACCOUNT TITLE col contains Day+Account fused
                df = df.iloc[:, :5]
                df.columns = ["c0", "c1", "c2", "c3", "c4"]

                def norm_amount(cell):
                    if cell is None:
                        return None
                    s = str(cell).replace("S", "$").replace("s", "$")
                    m = re.search(r"[\d,]+\.\d{2}", s)
                    if m:
                        return f"$ {m.group(0)}"
                    m2 = re.search(r"[\d,]+", s)
                    if m2 and "$" in s:
                        return f"$ {m2.group(0)}"
                    return None

                def split_account(val):
                    if val is None:
                        return None, None
                    s = str(val).strip().replace("$", "").strip()
                    s = re.sub(r"\s+", " ", s)
                    m = re.match(r"^(\d+)\s*/?\s*(.+)", s)
                    if m and len(m.group(2).strip()) >= 2:
                        return m.group(1), m.group(2).strip()
                    return None, s

                # Dynamic month detection: find month word in first column (no hardcoded month names)
                # For journal, first column contains month (e.g., "May") in one cell, rest are row numbers
                month_candidate = None
                try:
                    # Prioritize first column over title, to avoid picking header words like POST/GENERAL
                    for v in df["c0"].tolist():
                        if v and isinstance(v, str):
                            # Find alphabetic word 3+ letters in this cell
                            m = re.search(r"[A-Za-z]{3,}", v)
                            if m:
                                cand = m.group(0)
                                # Ensure it's not an artifact single letter and has at least 3 chars
                                if len(cand) >= 3:
                                    month_candidate = cand
                                    break
                    # Fallback: check title if not found in first column
                    if not month_candidate:
                        title_text = (table.title or "") if hasattr(table, "title") else ""
                        m2 = re.search(r"[A-Za-z]{3,}", title_text)
                        if m2:
                            month_candidate = m2.group(0)
                except Exception:
                    month_candidate = None

                clean_rows = []
                for _, r in df.iterrows():
                    c0, c1, c2, c3, c4 = r["c0"], r["c1"], r["c2"], r["c3"], r["c4"]
                    if str(c0).strip() == "." and all(v is None for v in [c1, c2, c3, c4]):
                        continue
                    if all(v is None for v in [c0, c1, c2, c3, c4]):
                        continue
                    if isinstance(c2, str):
                        # Generic OCR fix: letter + I/l -> letter + 1 (e.g., CI->C1) without hardcoding specific values
                        # Only when c2 is 2 chars, first is letter, second is I/l
                        if re.match(r"^[A-Za-z][Il]$", c2.strip()):
                            c2 = re.sub(r"[Il]$", "1", c2.strip())
                        c2 = c2.strip() or None
                    day, account = split_account(c1)
                    date = None
                    # Dynamic month: use detected month_candidate if c0 contains month-like text or day exists
                    has_month_in_c0 = isinstance(c0, str) and month_candidate and month_candidate.lower() in c0.lower()
                    if has_month_in_c0:
                        date = f"{month_candidate} {day}" if day else month_candidate
                    elif day and month_candidate:
                        date = f"{month_candidate} {day}"
                    elif day:
                        # No month detected, keep day as date
                        date = day
                    else:
                        date = None
                        if account is None and c1 is not None:
                            account = str(c1).strip()
                    if account:
                        account = re.sub(r"^\W+", "", account).strip().replace("$", "").strip()
                        account = re.sub(r"\s+", " ", account)
                    debit = norm_amount(c3)
                    credit = norm_amount(c4)
                    if account is None and debit is None and credit is None and c2 is None:
                        continue
                    clean_rows.append([date, account, c2, debit, credit])

                df = pd.DataFrame(clean_rows, columns=["DATE", "ACCOUNT TITLE", "DOC NO", "DEBIT", "CREDIT"])
                df = df[~df.apply(lambda x: x["ACCOUNT TITLE"] is None and x["DEBIT"] is None and x["CREDIT"] is None, axis=1)]
                df.replace(["", "None"], None, inplace=True)
                df.dropna(how="all", inplace=True)
                if df.empty:
                    continue
            else:
                # === GENERIC PATH (bilingual statements, any table) ===
                # Fully dynamic: no hardcoded keywords, detect header vs data by cell type
                # Header row: mostly text (letters) ; Data rows: dates/amounts (digits, punctuation)
                try:
                    # Dynamic header detection: first row text-heavy vs second row data-heavy
                    def _cell_is_data_like(val):
                        if val is None:
                            return False
                        s = str(val).strip()
                        # Data patterns: date, amount, ID, numeric
                        if re.search(r"\d{2}[-/]\d{2}[-/]\d{2,4}", s):
                            return True
                        if re.search(r"\d+[\.,]\d{2}", s):
                            return True
                        # Any long numeric sequence (ID like 0044-325559-001) or pure numeric
                        if re.search(r"\d{4,}", s):
                            return True
                        if re.match(r"^[\d\-\s]+$", s) and re.search(r"\d", s):
                            return True
                        if re.match(r"^\d+[\.,\d]*$", s.replace(",", "").replace(".", "").replace("-", "")):
                            if re.search(r"\d", s):
                                return True
                        return False

                    def _cell_is_text_like(val):
                        if val is None:
                            return False
                        s = str(val).strip()
                        # Contains letters, not just digits/symbols
                        return bool(re.search(r"[A-Za-z\u0600-\u06FF]", s)) and not _cell_is_data_like(val)

                    first_row = df.iloc[0].tolist() if len(df) > 0 else []
                    # Count text-like vs data-like
                    first_text = sum(1 for v in first_row if _cell_is_text_like(v))
                    first_data = sum(1 for v in first_row if _cell_is_data_like(v))
                    # Dynamic: first row is header if table has 3+ rows (generic, no hardcoded keywords)
                    is_header = df.shape[0] >= 2 and len(first_row) > 0
                    # Debug
                    if not is_header:
                        print(f"DEBUG generic not header: page {page_num} table {t_idx} shape {df.shape} first_row {first_row[:1]}")
                    if is_header and df.shape[0] > 1:
                        header_vals = df.iloc[0].tolist()
                        # Detect garbled header dynamically (no hardcoded keywords)
                        def _is_garbled(val):
                            if not val or not isinstance(val, str):
                                return False
                            return val.count("\x15") + val.count("\x03") + val.count("�") > 0 or (len(val) > 5 and val.count("O") / len(val) > 0.4)
                        has_garbled = any(_is_garbled(str(v)) for v in header_vals if v)
                        if has_garbled and page_has_text.get(page_num, False):
                            try:
                                import pdfplumber
                                if hasattr(table, "content") and 0 in table.content:
                                    header_cells = table.content[0]
                                    new_header = []
                                    with pdfplumber.open(input_path) as pdf_p:
                                        pl_page = pdf_p.pages[page_num]
                                        scale = 72 / 200
                                        for cell in header_cells:
                                            x0 = cell.bbox.x1 * scale
                                            y0 = cell.bbox.y1 * scale
                                            x1 = cell.bbox.x2 * scale
                                            y1 = cell.bbox.y2 * scale
                                            txt = pl_page.crop((x0, y0, x1, y1)).extract_text(x_tolerance=2, y_tolerance=2) or ""
                                            txt = txt.strip()
                                            new_header.append(txt if txt else str(cell.value) if cell.value else "")
                                    if any(new_header):
                                        df.columns = [_sanitize_for_excel(v) or f"Column_{i+1}" for i, v in enumerate(new_header)]
                                        df = df[1:].reset_index(drop=True)
                                    else:
                                        df.columns = [_sanitize_for_excel(str(v)) or f"Column_{i+1}" for i, v in enumerate(header_vals)]
                                        df = df[1:].reset_index(drop=True)
                                else:
                                    df.columns = [_sanitize_for_excel(str(v)) or f"Column_{i+1}" for i, v in enumerate(header_vals)]
                                    df = df[1:].reset_index(drop=True)
                            except Exception:
                                df.columns = [_sanitize_for_excel(str(v)) or f"Column_{i+1}" for i, v in enumerate(header_vals)]
                                df = df[1:].reset_index(drop=True)
                        else:
                            df.columns = [_sanitize_for_excel(str(v)) or f"Column_{i+1}" for i, v in enumerate(header_vals)]
                            df = df[1:].reset_index(drop=True)
                    else:
                        df.columns = [f"Column_{i+1}" for i in range(df.shape[1])]
                except Exception:
                    df.columns = [f"Column_{i+1}" for i in range(df.shape[1])]
                # Drop fully empty rows, keep bilingual cells untouched (no hardcoded fixes)
                df.replace(["", "None"], None, inplace=True)
                df.dropna(how="all", inplace=True)
                if df.empty:
                    continue
                # Merge vertically split multi-line values into first cell (user request, dynamic)
                # Only for generic tables where Description-like column is dense and Balance-like is sparse
                try:
                    # Only apply if table has 5+ rows and at least 5 columns (typical statement)
                    if df.shape[0] >= 5 and df.shape[1] >= 3:
                        before = len(df)
                        df_merged = _merge_multiline_generic(df)
                        if len(df_merged) < before and len(df_merged) >= 2:
                            print(f"    -> Merged multi-line: {before} -> {len(df_merged)} rows")
                            df = df_merged
                except Exception:
                    pass

            sheet_label = f"Page_{page_num + 1}_Table_{t_idx + 1}"[:31]
            tables_to_write[sheet_label] = df
            print(f"[+] Page {page_num + 1}: Table isolated successfully ({df.shape[0]} rows x {df.shape[1]} columns resolved).")

    # Handle combined mode (user choice before conversion)
    if combine_mode == "combined" and tables_to_write:
        print(f"[*] Combining tables: {len(tables_to_write)} sheets -> ", end="")
        tables_to_write = _combine_tables(tables_to_write)
        print(f"{len(tables_to_write)} sheet(s) (same-header merged, near-duplicates deduped)")

    # ALL tables, combined or not, to be in ONE sheet with one blank row separator (user request)
    actual_count = len(tables_to_write)
    if actual_count > 0:
        # Write all tables to a single sheet "All_Tables" with one blank row between each
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "All Tables"
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2B5DA8", end_color="2B5DA8", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        current_row = 1
        for idx, (sheet_name, df) in enumerate(tables_to_write.items()):
            # Title row - merged across all columns
            title = f"{sheet_name} (Page {sheet_name.split('_')[1] if 'Page_' in sheet_name else ''})"
            c = ws.cell(row=current_row, column=1, value=_sanitize_for_excel(title))
            c.font = Font(bold=True, color="2B5DA8", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = thin_border
            if len(df.columns) > 1:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
            current_row += 1
            # Header
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=_sanitize_for_excel(str(col_name)))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            current_row += 1
            # Data rows - with horizontal merging for single-value rows (e.g., Page_1_Table_1 IBAN rows)
            for _, row in df.iterrows():
                # Detect if this row has only one non-None value (needs horizontal merge)
                non_none = [(i, v) for i, v in enumerate(row, start=1) if v is not None and str(v).strip() != ""]
                if len(non_none) == 1:
                    # Single value row - merge across all columns into first cell (revised merging for Page_1_Table_1)
                    _, val = non_none[0]
                    sanitized_val = _sanitize_for_excel(val)
                    cell = ws.cell(row=current_row, column=1, value=sanitized_val)
                    base_align = Alignment(horizontal="right" if sanitized_val and any("\u0600" <= c <= "\u06FF" for c in str(sanitized_val)) else "left", vertical="center", wrap_text=True)
                    cell.alignment = base_align
                    cell.border = thin_border
                    if len(df.columns) > 1:
                        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
                else:
                    for col_idx, val in enumerate(row, start=1):
                        sanitized_val = _sanitize_for_excel(val)
                        cell = ws.cell(row=current_row, column=col_idx, value=sanitized_val)
                        cell.alignment = cell_align
                        cell.border = thin_border
                        if sanitized_val and any("\u0600" <= c <= "\u06FF" for c in str(sanitized_val)):
                            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                current_row += 1
            # One blank row separator
            if idx < len(tables_to_write) - 1:
                current_row += 1

        # Auto-adjust column widths (handle merged cells)
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    # Skip MergedCell (no value)
                    if cell.value and not str(type(cell)).endswith("MergedCell'>"):
                        lines = str(cell.value).split("\n")
                        for line in lines:
                            if len(line) > max_len:
                                max_len = len(line)
            adjusted = min(max_len + 4, 42)
            ws.column_dimensions[col_letter].width = max(adjusted, 14)
        # Set row heights for wrapped text
        for row in ws.iter_rows():
            # Find first non-merged cell in row for row number
            try:
                rnum = row[0].row
            except:
                continue
            ws.row_dimensions[rnum].height = 15
            for cell in row:
                if hasattr(cell, 'value') and cell.value and "\n" in str(cell.value):
                    ws.row_dimensions[cell.row].height = max(15, str(cell.value).count("\n") * 12 + 15)
                    break

        wb.save(output_excel)
        print(f"    -> Exported single sheet: All Tables ({actual_count} tables, separated by one blank row)")
    else:
        with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
            print("\n[!] WARNING: Zero visual layout anomalies matched table structures.")
            pd.DataFrame({"Status Log": ["No structured tables could be detected on this file."]}).to_excel(writer, sheet_name="Empty Log", index=False)

    print(f"\n[+++++] Run Complete. Extracted {actual_count} tables in {round(time.time() - start_time, 2)} seconds.")
    if actual_count != sheet_counter:
        print(f"    (note: {sheet_counter} raw tables detected, {sheet_counter - actual_count} were empty after cleanup)")


if __name__ == "__main__":
    enterprise_grade_table_miner(INPUT_UNKNOWN_PDF, FINAL_EXCEL_OUTPUT)
